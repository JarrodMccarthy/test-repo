import xlrd
import json
from openpyxl import Workbook 
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd 

SavePath = 'C:/Jarrods/RodneyJob/WriteAPL.xlsx' #change for each sheet to write
Path1Json = 'C:/Jarrods/RodneyJob/PLCCodeExports/' 

PLCList = ['PLC20X','PLC21X', 'PLC30X', 'PLC31X', 'PLC40X', 'PLC41X', 'PLC50X', 'PLC51X', 'PLC52X', 'PLC60X', 'PLC61X', 'PLC71X'] #change for various PLCLIst
pathXL = 'C:/Jarrods/RodneyJob/APL.xlsx' #chnage for each xl sheet to read

paths = []

READWb = xlrd.open_workbook(pathXL)

WriteWb = Workbook()

sheet_names = READWb.sheet_names()

Write_Sheet_Names = []

TestTemplate = {'JFN': [], 'SED': [], 'SFD': [], 'XFD': [], 'XFN': [], 'EXL': [], 'PMPA': [], 'PMPB': [], 'PMPL': [], 'PMPS': []}

D_M_Map = {'JFN': 'JFN', 'SED': 'SDA', 'SFD': 'SFD', 'XFD': 'XFD', 'XFN': 'XFN', 'EXL': 'ECL', 'PMPA': 'SWP', 'PMPB': 'SWP', 'PMPL': 'SWP', 'PMPS': 'GWP'}

for ReadName in sheet_names:
	TestTemplate[ReadName] = pd.read_excel(pathXL, ReadName)
	Write_Sheet_Names.append(ReadName)

Write_Worksheet = {} #defines an empty dictionary of worksheets 
Write_Worksheet_Keys = [] #defines an empty list of worksheet keys

for WriteName in Write_Sheet_Names:
	Write_Worksheet_Keys.append(WriteName) #builds the worksheet keys list based on the Write_sheet_names
	

count = 0

for k in Write_Worksheet_Keys:
	Write_Worksheet[k] = WriteWb.create_sheet(k, count) # places the worksheets in the workbook created above in the order they are defined in inside the read workbook, with the name specified by the keys
	count = count+1

IO_Tag_Map = {}
IOcount = 0 

for sheet in sheet_names:
	currentsheet = READWb.sheet_by_name(sheet)
	headers = currentsheet.row_values(0)
	currentPLC = currentsheet.cell(1,1).value
	currentDevice = currentsheet.cell(1,2).value
	count = 0
	for row in range(1, currentsheet.nrows):
		if currentsheet.cell(row, 4).value == 1:
			count = count+1
		else:
			break

	NumberOfIO = count
	currentIO = []
	

	for row in range(1, NumberOfIO+1): #needs to be more sophisticated 
		currentIO.append(currentsheet.cell(row,7).value) #if it is not more sophisticated, change 7 to 8 when using latest TestTemplates
	
	
	parameters = currentIO
	
	for i in range(0,len(PLCList)):
		Fullpath = Path1Json+PLCList[i]+'.json'
		paths.append(Fullpath)
		

	for path in paths:
		split = path.split("/")[4]
		currentPLC = split.split(".")[0]
		
		
		with open(path, "r") as f: 
		    data = f.read() #returns the doc as a str

		dic = json.loads(data) #Returns the document as a dictionary 

		json_string = json.dumps(dic, indent = 4) #returns a formatted string

		DevicesInJSON = []
		ModulesInJSON = []
		FieldStuffInJSON = []
		IOInJSON = []
		OPCTags = []
		for device in dic.keys():
			DevicesInJSON.append(device)
			#print(device)
			for module in dic[device].keys():
				ModulesInJSON.append(module)
				#print(module)
				for field in dic[device][module].keys():
					#print(field)
					if field != 'DependantDevices' and field != 'HealthyBitsToLookup':
						FieldStuffInJSON.append(field)
						for IO in dic[device][module][field].keys():
							#print(IO)
							IOInJSON.append(IO)
							try:
								if D_M_Map[sheet] in IO:
									for parameter in parameters:
										if parameter in IO:
											
											
											Tag = f"ns=3;s=::[{currentPLC}]{dic[device][module][field][IO]}"

											IO_Tag_Map.update({IO:Tag})

											#print(f"{IO}, {Tag}")
									
							except:
								pass


PLC_Module_Map = {}

for key in IO_Tag_Map:
	module = key.split('.')[0]
	PLC = IO_Tag_Map[key].split('[')[1]
	PLC = PLC.split(']')[0]
	PLC_Module_Map.update({module:PLC})

Tests = {}

for module in PLC_Module_Map:
	for key in D_M_Map:
		if D_M_Map[key] in module:
			Tests.update({module: TestTemplate[key]})
			#print(type(TestTemplate[key]))

count = 0

for WriteName in Write_Sheet_Names:
		Write_Worksheet_Keys.append(WriteName) #builds the worksheet keys list based on the Write_sheet_names

for k in Write_Worksheet_Keys:
	Write_Worksheet[k] = WriteWb.create_sheet(k, count) # places the worksheets in the workbook created above in the order they are defined in inside the read workbook, with the name specified by the keys
	count = count+1
	#print(Write_Worksheet)
for module in Tests.keys():
	print(type(Tests[module]))
	print(f"{module}"'\n')
	print(Tests[module])

for module in Tests.keys():

	rowcount = len(Tests[module])

	
	for row in range(0, rowcount):
		try:
			IORef = module+'.'+Tests[module].iloc[row,7]
			
			Tests[module].iloc[row,1] = PLC_Module_Map[module]
			
			Tests[module].iloc[row,3] = module
			
			Tests[module].iloc[row,9] = IO_Tag_Map[IORef]
			
			
		except:
			pass

	#print(module) #write dataframe to a sheet right here...
	#print(Tests[module])


		
#		for key in Write_Worksheet.keys():
#			for rows in dataframe_to_rows(Tests[module], index = True, header = True):
#					Write_Worksheet[key].append(rows)
#
#		WriteWb.save(SavePath)

